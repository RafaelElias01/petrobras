"""
Abre _OFERTAS CORNEAS Nova Planilha.xlsx e acrescenta:
  - Dropdowns SNT nas colunas MR (D-G e J-M) da aba OFERTAS
  - Coluna N (EXPORTAR) — usuário digita X para marcar linha
  - Aba RECUSAS com cabeçalho pronto
  - Aba LISTAS oculta com os 39 motivos do SNT
Salva como OFERTAS_CORNEAS_NOVA.xlsx
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ORIGEM = "_OFERTAS CORNEAS Nova Planilha.xlsx"
DESTINO = "OFERTAS_CORNEAS_NOVA.xlsx"

MOTIVOS = [
    "Alteração laboratorial","Alteração morfológica","Alterações no tecido",
    "Antecedentes mórbidos","Cardiopatia - coronariopatia",
    "Cardiopatia - hipertensão arterial","Cardiopatia - miocardiopatia",
    "Cardiopatia - valvulopatia","Condições do Doador","Diabetes",
    "Distância","Droga vasopressora","Falta de cateterismo/eco","Idade",
    "Infecção","Instabilidade hemodinâmica","Lesão do órgão",
    "Má perfusão do órgão","Não ofertado","Preservação inadequada do órgão",
    "Ranking esgotado","SARS-CoV-2 Positivo","Sem equipe para transplante",
    "Sem receptores","Sorologia - Chagas","Sorologia - Hepatite B",
    "Sorologia - Hepatite C","Sorologia - HTLV I/II","Sorologia - Sífilis",
    "Sorologia - Toxoplasmose/Citomegalovirus","Sorologia não realizada",
    "Tamanho ou Peso","Tempo de isquemia fria",
    "Tempo prolongado de intubação/internação","Usuário de droga injetável",
    "Utilizado para pesquisa","Utilizado para transplante de ilhotas",
    "Utilizado para valvas cardíacas","Utilizado parente/cônjuge",
]

# estilos
def fill(h): return PatternFill("solid", fgColor=h)
thin  = Side(style="thin")
borda = Border(left=thin, right=thin, top=thin, bottom=thin)
centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Abrir original ────────────────────────────────────────────
wb    = openpyxl.load_workbook(ORIGEM)
ws_of = wb["OFERTAS"]

# ═══ 1. ABA LISTAS (oculta) ══════════════════════════════════
if "LISTAS" in wb.sheetnames:
    del wb["LISTAS"]
ws_l = wb.create_sheet("LISTAS")
ws_l.sheet_state = "hidden"
for i, v in enumerate(MOTIVOS, 1):
    ws_l.cell(i, 1, v)

n_mr = len(MOTIVOS)
ref_mr = f"LISTAS!$A$1:$A${n_mr}"

# ═══ 2. DROPDOWNS MR nas colunas D-G e J-M (linhas 3-15) ════
# Remove validações antigas nessas colunas
ws_of.data_validations.dataValidation = [
    dv for dv in ws_of.data_validations.dataValidation
    if not any(c in str(dv.sqref) for c in list("DEFGJKLM"))
]

def add_dv(ws, formula, sqref, titulo, msg):
    dv = DataValidation(type="list", formula1=formula,
                        allow_blank=True, showDropDown=False)
    dv.promptTitle, dv.prompt, dv.sqref = titulo, msg, sqref
    ws.add_data_validation(dv)

add_dv(ws_of, ref_mr, "D3:G15", "Motivo Recusa OD",
        "Selecione o motivo de recusa - Olho Direito")
add_dv(ws_of, ref_mr, "J3:M15", "Motivo Recusa OE",
        "Selecione o motivo de recusa - Olho Esquerdo")

# ═══ 3. COLUNA N — EXPORTAR ══════════════════════════════════
ws_of["N1"] = "→ Digite X para exportar"
ws_of["N1"].font      = Font(italic=True, size=8, color="C00000")
ws_of["N1"].alignment = Alignment(horizontal="center", wrap_text=True)

ws_of["N2"] = "EXPORTAR"
ws_of["N2"].fill      = fill("C00000")
ws_of["N2"].font      = Font(color="FFFFFF", bold=True, size=10)
ws_of["N2"].alignment = centro
ws_of["N2"].border    = borda

for r in range(3, 16):
    c = ws_of.cell(r, 14)
    c.fill      = fill("FFCCCC")
    c.border    = borda
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.font      = Font(bold=True, size=12, color="C00000")

ws_of.column_dimensions["N"].width = 13

# ═══ 4. ABA RECUSAS ══════════════════════════════════════════
if "RECUSAS" in wb.sheetnames:
    del wb["RECUSAS"]
ws_r = wb.create_sheet("RECUSAS")
ws_r.sheet_properties.tabColor = "C00000"

ws_r.merge_cells("A1:O1")
c = ws_r["A1"]
c.value, c.fill, c.font, c.alignment = (
    "RECUSAS EXPORTADAS",
    fill("C00000"),
    Font(color="FFFFFF", bold=True, size=13),
    centro
)
ws_r.row_dimensions[1].height = 30

cabs = ["RGCT","OFERTA OD","HOSPITAL OD","MR-1 OD","MR-2 OD","MR-3 OD",
        "MR-4 OD","OFERTA OE","HOSPITAL OE","MR-1 OE","MR-2 OE","MR-3 OE",
        "MR-4 OE","—","DATA EXPORTAÇÃO"]
for ci, nome in enumerate(cabs, 1):
    c = ws_r.cell(2, ci)
    c.value, c.fill, c.font, c.alignment, c.border = (
        nome, fill("1F4E79"),
        Font(color="FFFFFF", bold=True), centro, borda
    )
ws_r.row_dimensions[2].height = 20

larguras = [14,25,13,28,28,28,28,25,13,28,28,28,28,5,18]
from openpyxl.utils import get_column_letter
for ci, larg in enumerate(larguras, 1):
    ws_r.column_dimensions[get_column_letter(ci)].width = larg

# ═══ Ordenar abas ════════════════════════════════════════════
idx_of = wb.sheetnames.index("OFERTAS")
wb.move_sheet("RECUSAS", offset=idx_of + 1 - wb.sheetnames.index("RECUSAS"))
wb.move_sheet("LISTAS",  offset=len(wb.sheetnames) - 1 - wb.sheetnames.index("LISTAS"))

wb.active = wb["OFERTAS"]
wb.save(DESTINO)

print(f"✔ Salvo: {DESTINO}")
print(f"  - Dropdowns MR em D3:G15 e J3:M15 ({n_mr} motivos SNT)")
print(f"  - Coluna N (EXPORTAR): digit X na linha → rode exportar_recusas.py")
print(f"  - Aba RECUSAS criada")
print(f"  - Aba LISTAS oculta")
