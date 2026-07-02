"""
EXPORTAR RECUSAS — executa como o "botão" da planilha.

Fluxo de uso:
  1. Na aba OFERTAS, coluna N (EXPORTAR), digite  X  na linha que quer exportar.
  2. Salve e feche o Excel.
  3. Dê duplo clique em EXPORTAR.bat (ou: python exportar_recusas.py)
  4. A linha marcada vai para a aba RECUSAS e a marcação X é apagada.
"""
import sys
import os
import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

ARQUIVO      = "OFERTAS_CORNEAS_NOVA.xlsx"
ABA_OF       = "OFERTAS"
ABA_REC      = "RECUSAS"
LINHA_INI    = 3
LINHA_FIM    = 15
COL_EXPORTAR = 14          # coluna N = 14
COLUNAS_MR   = [4,5,6,7,10,11,12,13]

# ─── helpers de estilo ───────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
thin  = Side(style="thin")
borda = Border(left=thin, right=thin, top=thin, bottom=thin)

# ─── Verificar arquivo ───────────────────────────────────────
if not os.path.exists(ARQUIVO):
    print(f"ERRO: '{ARQUIVO}' não encontrado na pasta:")
    print(f"  {os.getcwd()}")
    input("\nPressione Enter para fechar...")
    sys.exit(1)

print(f"Abrindo {ARQUIVO}...")
wb = openpyxl.load_workbook(ARQUIVO)

if ABA_OF not in wb.sheetnames:
    print(f"ERRO: aba '{ABA_OF}' não encontrada.")
    input("\nPressione Enter para fechar...")
    sys.exit(1)

ws_of = wb[ABA_OF]

# ─── Criar / garantir aba RECUSAS ────────────────────────────
if ABA_REC not in wb.sheetnames:
    ws_rec = wb.create_sheet(ABA_REC)
    ws_rec.sheet_properties.tabColor = "C00000"
else:
    ws_rec = wb[ABA_REC]

# Cabeçalho da aba RECUSAS (cria só se vazio)
if ws_rec.cell(1, 1).value is None:
    ws_rec.merge_cells("A1:O1")
    c = ws_rec["A1"]
    c.value     = "RECUSAS EXPORTADAS"
    c.fill      = fill("C00000")
    c.font      = Font(color="FFFFFF", bold=True, size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_rec.row_dimensions[1].height = 28

if ws_rec.cell(2, 1).value is None:
    cabecalhos = [
        "RGCT", "OFERTA OD", "HOSPITAL OD",
        "MR-1 OD", "MR-2 OD", "MR-3 OD", "MR-4 OD",
        "OFERTA OE", "HOSPITAL OE",
        "MR-1 OE", "MR-2 OE", "MR-3 OE", "MR-4 OE",
        "—",           # col N (EXPORTAR) — ignorada
        "DATA EXPORTAÇÃO"
    ]
    for ci, nome in enumerate(cabecalhos, 1):
        c = ws_rec.cell(2, ci)
        c.value     = nome
        c.fill      = fill("1F4E79")
        c.font      = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borda

# ─── Próxima linha livre em RECUSAS ──────────────────────────
proxima_rec = 3
for r in range(ws_rec.max_row, 2, -1):
    if any(ws_rec.cell(r, c).value for c in range(1, 16)):
        proxima_rec = r + 1
        break

# ─── Procurar linhas marcadas com X ──────────────────────────
agora      = datetime.now().strftime("%d/%m/%Y %H:%M")
exportadas = []

for linha in range(LINHA_INI, LINHA_FIM + 1):
    marcador = ws_of.cell(linha, COL_EXPORTAR).value
    if marcador is None:
        continue
    if str(marcador).strip().upper() not in ("X", "SIM", "S", "1", "EXPORTAR"):
        continue

    # Copiar A:M para RECUSAS
    for col in range(1, 14):
        destino       = ws_rec.cell(proxima_rec, col)
        destino.value = ws_of.cell(linha, col).value

    # Data na coluna O (15)
    ws_rec.cell(proxima_rec, 15).value = agora

    # Apagar MRs e marcador na linha original
    for col in COLUNAS_MR:
        ws_of.cell(linha, col).value = None
    ws_of.cell(linha, COL_EXPORTAR).value = None

    exportadas.append(linha)
    proxima_rec += 1

# ─── Salvar ──────────────────────────────────────────────────
wb.save(ARQUIVO)

# ─── Resultado ───────────────────────────────────────────────
print()
if not exportadas:
    print("Nenhuma linha marcada para exportar.")
    print("  → Digite X na coluna N (EXPORTAR) da linha desejada,")
    print("    salve o Excel e rode este script novamente.")
else:
    print(f"✔ {len(exportadas)} recusa(s) exportada(s):")
    for l in exportadas:
        rgct = wb[ABA_REC].cell(proxima_rec - len(exportadas) + exportadas.index(l), 1).value or "?"
        print(f"   Linha {l} — RGCT: {rgct}")
    print(f"\n  MRs apagados da aba '{ABA_OF}'.")
    print(f"  Dados salvos na aba '{ABA_REC}'.")
    print(f"  Arquivo: {ARQUIVO}")

input("\nPressione Enter para fechar...")
