import openpyxl
wb = openpyxl.load_workbook("_OFERTAS CORNEAS Nova Planilha.xlsx", data_only=True)
ws = wb["OFERTAS"]
for r in range(1, 30):
    vals = [(ws.cell(r,c).coordinate, ws.cell(r,c).value) for c in range(1, ws.max_column+1) if ws.cell(r,c).value is not None]
    if vals:
        print(f"R{r}: {vals}")
print()
ws2 = wb["EQUIPES"]
for r in range(1, 40):
    vals = [(ws2.cell(r,c).coordinate, ws2.cell(r,c).value) for c in range(1, ws2.max_column+1) if ws2.cell(r,c).value is not None]
    if vals:
        print(f"R{r}: {vals}")
print()
ws3 = wb["DESCARTES"]
for r in range(1, 30):
    vals = [(ws3.cell(r,c).coordinate, ws3.cell(r,c).value) for c in range(1, ws3.max_column+1) if ws3.cell(r,c).value is not None]
    if vals:
        print(f"R{r}: {vals}")
